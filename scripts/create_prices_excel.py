"""Create a formatted Excel file for price research — populated from resorts.json."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEED    = Path(__file__).parent.parent / "data" / "seed" / "resorts.json"
OUTPUT  = Path(__file__).parent.parent / "data" / "seed" / "prices_research.xlsx"

COUNTRY_HE = {
    "FR": "צרפת", "CH": "שווייץ", "AT": "אוסטריה", "IT": "איטליה",
    "SE": "שוודיה", "NO": "נורבגיה", "AD": "אנדורה", "ES": "ספרד",
    "BG": "בולגריה", "SK": "סלובקיה", "PL": "פולין",
    "DE": "גרמניה", "CZ": "צ'כיה", "RO": "רומניה", "SI": "סלובניה", "LI": "ליכטנשטיין",
}

TIER_HE = {"premium": "פרמיום", "mid": "בינוני", "budget": "זול"}

HEADERS = [
    "שם הריזורט", "מדינה", "אתר רשמי", "עונה",
    "כרטיס יומי מבוגר שיא (€)", "כרטיס יומי ילד שיא (€)",
    "כרטיס 6 ימים מבוגר (€)", "כרטיס 6 ימים ילד (€)",
    "שעות פתיחה", "חודשי שיא", "חודשים שקטים",
    "השכרת סקי ליום (€)", "השכרת סנובורד ליום (€)", "קסדה ליום (€)",
    "דרגת מחיר", "קישור לכרטיסים", "הערות",
]

COL_WIDTHS = [32, 12, 40, 14, 24, 24, 24, 22, 16, 18, 18, 20, 22, 14, 14, 42, 30]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
PRICED_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green tint — has price data
EMPTY_FILL   = PatternFill("solid", fgColor="FCE4D6")   # orange tint — missing prices
ALT_PRICED   = PatternFill("solid", fgColor="C6EFCE")
ALT_EMPTY    = PatternFill("solid", fgColor="F8CBAD")
LINK_FONT    = Font(color="0563C1", underline="single", size=10)
BORDER       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

resorts = json.loads(SEED.read_text(encoding="utf-8"))

# Only the resorts that have price data — sorted by country then name
resorts_filtered = [r for r in resorts if r.get("adult_day_pass_peak_eur")]
resorts_filtered.sort(key=lambda r: (r["country"], r["name"]))

wb = Workbook()
ws = wb.active
ws.title = "מחירי ריזורטים"
ws.sheet_view.rightToLeft = True

# Header
for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.border    = BORDER
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 36

# Data rows
for row_idx, r in enumerate(resorts_filtered, 2):
    has_price = bool(r.get("adult_day_pass_peak_eur"))
    even = row_idx % 2 == 0
    if has_price:
        fill = ALT_PRICED if even else PRICED_FILL
    else:
        fill = ALT_EMPTY if even else EMPTY_FILL

    country_he = COUNTRY_HE.get(r["country"], r["country"])
    tier_he    = TIER_HE.get(r.get("price_tier", ""), "")
    website    = r.get("website_url", r.get("ticket_url", ""))
    season_label = r.get("price_season", "")

    data = [
        r["name"],                                      # 1  שם
        country_he,                                     # 2  מדינה
        website,                                        # 3  אתר
        season_label,                                   # 4  עונה
        r.get("adult_day_pass_peak_eur", ""),           # 5  מבוגר יום
        r.get("child_day_pass_peak_eur", ""),           # 6  ילד יום
        r.get("adult_6day_pass_eur", ""),               # 7  מבוגר 6 ימים
        r.get("child_6day_pass_eur", ""),               # 8  ילד 6 ימים
        r.get("opening_hours", ""),                     # 9  שעות
        r.get("peak_months", ""),                       # 10 שיא
        r.get("quiet_months", ""),                      # 11 שקט
        r.get("ski_rental_day_eur", ""),                # 12 סקי
        r.get("snowboard_rental_day_eur", ""),          # 13 סנובורד
        r.get("helmet_rental_day_eur", ""),             # 14 קסדה
        tier_he,                                        # 15 דרגה
        r.get("ticket_url", ""),                        # 16 קישור כרטיסים
        "",                                             # 17 הערות (ריק לעריכה)
    ]

    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = BORDER
        cell.fill   = fill
        if col in (3, 16):  # URL columns
            cell.font      = LINK_FONT
            cell.alignment = CENTER
        elif col in (5, 6, 7, 8, 12, 13, 14):  # price columns
            cell.alignment = CENTER
            if value != "":
                cell.number_format = "#,##0.00"
        elif col in (9, 10, 11):  # hours / months
            cell.alignment = CENTER
        else:
            cell.alignment = RIGHT

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# Legend sheet
ws2 = wb.create_sheet("מקרא")
ws2.sheet_view.rightToLeft = True
legend = [
    ("צבע", "משמעות"),
    ("ירוק", "מחיר מלא — מוכן"),
    ("כתום", "חסר מחיר — יש להשלים"),
    ("", ""),
    ("דרגת מחיר", ""),
    ("פרמיום", "מעל 60€ ליום"),
    ("בינוני", "40-60€ ליום"),
    ("זול", "מתחת ל-40€ ליום"),
]
for row_idx, (a, b) in enumerate(legend, 1):
    ws2.cell(row=row_idx, column=1, value=a)
    ws2.cell(row=row_idx, column=2, value=b)

wb.save(OUTPUT)

priced_count  = sum(1 for r in resorts_filtered if r.get("adult_day_pass_peak_eur"))
missing_count = len(resorts_filtered) - priced_count
print(f"Excel created: {OUTPUT}")
print(f"  Priced (green):  {priced_count}")
print(f"  Missing (orange): {missing_count}")
print(f"  Total rows: {len(resorts_filtered)}")